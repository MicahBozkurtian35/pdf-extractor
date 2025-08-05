import anthropic
import base64
import os
import io
import xlwings as xw
import math
import pandas as pd
import openpyxl
import shutil
from pdf2image import convert_from_path
from PIL import Image
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import PieChart, Reference
import cv2
import numpy as np
from openpyxl.utils import get_column_letter
import pdfplumber
import json
import time
import glob
import requests  # Needed for Upstage API calls
import re

# Global debug flag: set to True for verbose output.
DEBUG = True
def debug_print(*args, **kwargs):
    if DEBUG:
        print(*args, **kwargs)

#############################################
# Helper: Remove Temporary Excel Files
#############################################
def remove_temp_excel_files(excel_path):
    """
    Removes any ~ prefixed Excel lock files from the directory containing excel_path.
    """
    directory = os.path.dirname(os.path.abspath(excel_path))
    temp_files = glob.glob(os.path.join(directory, "~$*"))
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
            debug_print("Removed temporary Excel file:", temp_file)
        except Exception as e:
            debug_print("Failed to remove temporary Excel file:", temp_file, e)

#############################################
# Helper: Wait for File Lock Release
#############################################
def wait_for_file_lock_release(file_path, timeout=30):
    """
    Repeatedly attempts to rename the file to itself to check if it's locked.
    Waits up to 'timeout' seconds for the file to be unlocked.
    """
    start_time = time.time()
    while True:
        try:
            os.rename(file_path, file_path)  # Attempt to rename
            break  # Success => not locked
        except Exception:
            time.sleep(1)
            if time.time() - start_time > timeout:
                debug_print("Timeout waiting for file lock release on", file_path)
                break

#############################################
# Basic Functions
#############################################
def read_anthropic_key(file_path=r"C:\API\anthropic_key.txt"):
    try:
        with open(file_path, "r") as f:
            return f.read().strip()
    except Exception as e:
        print(f"Error reading Anthropic API key from {file_path}: {e}")
        return None

def read_upstage_api_key(file_path=r"C:\API\upstage_api_key.txt"):
    """
    Reads your Upstage API key from a file, so you don't hardcode it.
    """
    try:
        with open(file_path, "r") as f:
            return f.read().strip()
    except Exception as e:
        print(f"Error reading Upstage API key from {file_path}: {e}")
        return None

def extract_chart_adaptive(pdf_path, page_num, bbox, output_path, add_margin=10):
    """
    Extracts a chart region from PDF using exact coordinates from Upstage with adaptive extensions
    to capture titles and labels.

    Args:
        pdf_path: Path to the PDF file
        page_num: Page number (1-based)
        bbox: Bounding box from Upstage (left, top, width, height)
        output_path: Path to save the extracted image
        add_margin: Additional margin to add around the chart

    Returns:
        Path to the extracted image file
    """
    import fitz  # PyMuPDF

    try:
        # Open the PDF document
        doc = fitz.open(pdf_path)

        # Convert from 1-based to 0-based page numbering
        page_idx = page_num - 1
        if page_idx < 0 or page_idx >= len(doc):
            debug_print(f"Page {page_num} out of range (document has {len(doc)} pages)")
            return None

        page = doc[page_idx]

        # Get PDF page dimensions
        page_width = page.rect.width
        page_height = page.rect.height
        debug_print(f"PDF page dimensions: {page_width} x {page_height}")

        # Get original bbox coordinates directly from Upstage (no scaling)
        left = bbox.get("left", 0)
        top = bbox.get("top", 0)
        width = bbox.get("width", 0)
        height = bbox.get("height", 0)

        debug_print(f"Original bbox: left={left}, top={top}, width={width}, height={height}")

        # Apply extensions to capture titles (above) and source notes/labels (below)
        top_extension = 75    # Extend upward to capture titles
        bottom_extension = 40  # Extend downward to capture source notes
        side_extension = 10    # Add small padding on sides

        # Create extended rectangle
        extended_rect = fitz.Rect(
            max(0, left - side_extension),                 # Left (don't go below 0)
            max(0, top - top_extension),                   # Top (don't go below 0)
            min(page_width, left + width + side_extension),  # Right (don't exceed page width)
            min(page_height, top + height + bottom_extension) # Bottom (don't exceed page height)
        )

        debug_print(f"Extended rectangle: {extended_rect}")
        debug_print(f"Top extension: {top - extended_rect.y0}px, Bottom extension: {extended_rect.y1 - (top + height)}px")

        # Extract the image with higher resolution
        pix = page.get_pixmap(clip=extended_rect, matrix=fitz.Matrix(2, 2))  # 2x scaling for better resolution
        pix.save(output_path)

        # Also save a debug version showing the extraction region
        debug_dir = os.path.dirname(output_path)
        debug_basename = os.path.basename(output_path)
        debug_path = os.path.join(debug_dir, f"debug_{debug_basename}")

        # Create a copy of the page for debugging
        debug_doc = fitz.open()
        debug_page = debug_doc.new_page(width=page_width, height=page_height)
        debug_page.show_pdf_page(debug_page.rect, doc, page_idx)

        # Draw the original rectangle in blue
        original_rect = fitz.Rect(left, top, left + width, top + height)
        debug_page.draw_rect(original_rect, color=(0, 0, 1), width=1.5)
        debug_page.insert_text(
            (left + 5, top + 15),
            "Original",
            fontsize=8,
            color=(0, 0, 1)
        )

        # Draw the extended rectangle in red
        debug_page.draw_rect(extended_rect, color=(1, 0, 0), width=2)
        debug_page.insert_text(
            (extended_rect.x0 + 5, extended_rect.y0 + 30),
            "Extended",
            fontsize=8,
            color=(1, 0, 0)
        )

        # Add information about extensions
        debug_page.insert_text(
            (20, 20),
            f"Chart extraction: Top +{top_extension}px, Bottom +{bottom_extension}px, Sides +{side_extension}px",
            fontsize=10,
            color=(0, 0, 0)
        )

        # Save the debug visualization
        debug_pix = debug_page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
        debug_pix.save(debug_path)
        debug_doc.close()

        debug_print(f"Extracted chart from PDF page {page_num} to {output_path}")
        debug_print(f"Debug image saved to {debug_path}")

        return output_path

    except Exception as e:
        debug_print(f"Error extracting chart from PDF: {e}")
        return None
    finally:
        if 'doc' in locals():
            doc.close()


def process_upstage_output(upstage_response, pdf_path=None):
    """
    Processes the Upstage API response to extract figure objects with their bounding boxes.
    Uses exact coordinates without any scaling transformations.
    """
    if not upstage_response:
        return []

    debug_print("Upstage response keys:", upstage_response.keys())
    figures = []

    # Check if the response contains the full HTML content
    if "content" in upstage_response and "html" in upstage_response["content"]:
        html_content = upstage_response["content"]["html"]

        # Try to extract page dimensions from HTML if available (for informational purposes only)
        page_dims_pattern = r'<div class="page" data-width="(\d+)" data-height="(\d+)"'
        page_dims_match = re.search(page_dims_pattern, html_content)
        page_width = int(page_dims_match.group(1)) if page_dims_match else None
        page_height = int(page_dims_match.group(2)) if page_dims_match else None

        if page_width and page_height:
            debug_print(f"Page dimensions from HTML: {page_width} x {page_height}")

        # Extract figure elements from HTML using regex
        figure_pattern = r'<figure id=[\'"](.*?)[\'"] data-category=[\'"](chart|table|figure)[\'"]><img data-coord="top-left:\((\d+),(\d+)\); bottom-right:\((\d+),(\d+)\)"'

        matches = re.findall(figure_pattern, html_content)
        debug_print(f"Found {len(matches)} figure matches in HTML")

        for i, match in enumerate(matches):
            fig_id, fig_type, left, top, right, bottom = match

            # Convert to integers
            left = int(left)
            top = int(top)
            right = int(right)
            bottom = int(bottom)
            width = right - left
            height = bottom - top

            # Skip tables - they're not charts we want to process
            if fig_type.lower() == 'table':
                debug_print(f"Skipping table figure with ID {fig_id}")
                continue

            # Use exact coordinates without any scaling
            figure = {
                "figure_id": fig_id,
                "page": 1,  # Default to page 1 if not specified
                "bounding_box": {
                    "left": left,
                    "top": top,
                    "width": width,
                    "height": height,
                    "right": right,
                    "bottom": bottom
                },
                "figure_type": fig_type,
                "original_order": i  # Keep track of original order for debugging
            }

            figures.append(figure)
            debug_print(f"Added figure {i+1}: ID={fig_id}, type={fig_type}, bbox={figure['bounding_box']}")

        # Sort figures by their position on the page (top to bottom, left to right)
        figures.sort(key=lambda x: (x["bounding_box"]["top"], x["bounding_box"]["left"]))

        # Update the figure indices after sorting
        for i, fig in enumerate(figures):
            debug_print(f"Sorted figure {i+1}: ID={fig['figure_id']}, original position {fig['original_order']+1}, top={fig['bounding_box']['top']}")

    # If we have a PDF path, create visualization with the exact coordinates
    if pdf_path and figures:
        visualize_pdf_with_bounding_boxes(pdf_path, figures)

    return figures


def visualize_pdf_with_bounding_boxes(pdf_path, figures, output_path="debug_pdf_boxes.pdf"):
    """
    Creates a new PDF with bounding boxes drawn around all figures.
    Shows both original and extended boundaries.
    """
    import fitz  # PyMuPDF

    try:
        # Open the source PDF
        doc = fitz.open(pdf_path)

        # For each page with figures
        for page_idx in range(len(doc)):
            page_num = page_idx + 1  # 1-based page number
            page = doc[page_idx]

            # Get page dimensions
            page_width = page.rect.width
            page_height = page.rect.height
            debug_print(f"Page {page_num} dimensions: {page_width} x {page_height}")

            # Find figures on this page
            page_figures = [f for f in figures if f.get("page", 1) == page_num]

            if not page_figures:
                continue

            debug_print(f"Drawing bounding boxes for {len(page_figures)} figures on page {page_num}")

            # Add a title
            page.insert_text(
                fitz.Point(50, 30),
                "Chart Detection & Extraction Visualization",
                fontsize=12,
                color=(0, 0, 0)
            )

            page.insert_text(
                fitz.Point(50, 50),
                "Blue: Original Upstage boundaries | Red: Extended boundaries (for extraction)",
                fontsize=10,
                color=(0, 0, 0)
            )

            # Draw extension parameters
            top_extension = 75
            bottom_extension = 40
            side_extension = 10

            page.insert_text(
                fitz.Point(50, 70),
                f"Extensions: {top_extension}px up, {bottom_extension}px down, {side_extension}px sides",
                fontsize=10,
                color=(0, 0, 0)
            )

            for fig_idx, fig in enumerate(page_figures):
                bbox = fig.get("bounding_box")
                if not bbox:
                    continue

                # Get original coordinates
                left = bbox.get("left", 0)
                top = bbox.get("top", 0)
                width = bbox.get("width", 0)
                height = bbox.get("height", 0)

                # Draw original rectangle (blue)
                original_rect = fitz.Rect(left, top, left + width, top + height)
                page.draw_rect(original_rect, color=(0, 0, 1), width=1.5)

                # Create and draw extended rectangle (red)
                extended_rect = fitz.Rect(
                    max(0, left - side_extension),
                    max(0, top - top_extension),
                    min(page_width, left + width + side_extension),
                    min(page_height, top + height + bottom_extension)
                )
                page.draw_rect(extended_rect, color=(1, 0, 0), width=1)

                # Add figure number and dimensions
                page.insert_text(
                    fitz.Point(left, top - 5),
                    f"Figure {fig_idx+1}",
                    fontsize=8,
                    color=(0, 0, 0)
                )

                page.insert_text(
                    fitz.Point(left, top + height + 15),
                    f"Original: {width}x{height}px",
                    fontsize=7,
                    color=(0, 0, 1)
                )

                extended_width = extended_rect.width
                extended_height = extended_rect.height
                page.insert_text(
                    fitz.Point(left, top + height + 25),
                    f"Extended: {extended_width:.0f}x{extended_height:.0f}px",
                    fontsize=7,
                    color=(1, 0, 0)
                )

        # Save the annotated PDF
        doc.save(output_path)
        doc.close()

        print(f"PDF with bounding boxes saved to {output_path}")
        return True

    except Exception as e:
        print(f"Error visualizing PDF with bounding boxes: {e}")
        import traceback
        traceback.print_exc()
        return False


def extract_chart_from_pdf(pdf_path, page_num, bbox, output_path, add_margin=10):
    """
    Legacy function maintained for compatibility.
    Now calls the new adaptive extraction function.
    """
    return extract_chart_adaptive(pdf_path, page_num, bbox, output_path, add_margin)

#############################################
# Upstage API Integration
#############################################
def extract_figures_with_upstage(pdf_path, api_key=None):
    """
    Sends the PDF to the Upstage Document Parser API and returns its JSON response.
    """
    if not api_key:
        api_key = read_upstage_api_key()

    if not api_key:
        print("Upstage API key not found.")
        return None

    # Corrected endpoint based on documentation
    url = "https://api.upstage.ai/v1/document-ai/document-parse"

    # Parameters based on the example
    params = {
        "ocr": "force",
        "base64_encoding": "['table']",
        "model": "document-parse"
    }

    headers = {
        "Authorization": f"Bearer {api_key}"
    }

    # PDF to send - use "document" as the field name from example
    files = {
        "document": ("document.pdf", open(pdf_path, "rb"), "application/pdf")
    }

    try:
        response = requests.post(
            url,
            headers=headers,
            files=files,
            data=params,
            timeout=30
        )

        if response.status_code != 200:
            print(f"Upstage API error: {response.status_code} - {response.text}")
            return None

        result = response.json()
        debug_print("Upstage response status code:", response.status_code)
        debug_print("Upstage response headers:", response.headers)
        debug_print("Upstage response sample:", result)
        return result
    except Exception as e:
        print("Error calling Upstage API:", e)
        return None
    finally:
        files["document"][1].close()  # Note the change from "file" to "document"

def filter_out_reference_lines(graph_data):
    """
    Filters out reference lines (Avg, SD, etc.) from graph data to keep only the main data series.
    """
    if not graph_data:
        return graph_data

    # Check if this is a chart with CSV data
    if "csv_data" not in graph_data:
        return graph_data

    csv_text = graph_data["csv_data"]

    # Parse the CSV to get the header row
    try:
        df = pd.read_csv(io.StringIO(csv_text))
        # Identify columns to keep (exclude Avg and SD columns)
        columns_to_keep = []
        for col in df.columns:
            # Keep columns that don't contain Avg, SD, average, or standard deviation references
            if not any(ref in col.lower() for ref in ["avg", "sd", "+1", "+2", "-1", "-2", "average", "deviation"]):
                columns_to_keep.append(col)

        # Only x-axis (first column) and the main data series
        if len(columns_to_keep) > 0:
            # Filter the DataFrame
            filtered_df = df[columns_to_keep]
            # Convert back to CSV string
            filtered_csv = filtered_df.to_csv(index=False)
            graph_data["csv_data"] = filtered_csv

            # Update data representation if it exists
            if "data_representation" in graph_data:
                # Update the data representation to reflect the kept columns
                for series_type in graph_data["data_representation"]:
                    if isinstance(graph_data["data_representation"][series_type], list):
                        filtered_series = []
                        for series_name in graph_data["data_representation"][series_type]:
                            if any(series_name.lower() in col.lower() for col in columns_to_keep):
                                filtered_series.append(series_name)
                        graph_data["data_representation"][series_type] = filtered_series
    except Exception as e:
        debug_print(f"Error filtering reference lines: {e}")

    return graph_data

#############################################
# Graph Detection with Claude
#############################################
def detect_graphs_with_claude(image_path, api_key):
    """
    Sends the image to Claude for analysis and extracts advanced chart data.
    """
    with open(image_path, "rb") as f:
        image_data = f.read()
    base64_encoded_image = base64.b64encode(image_data).decode("utf-8")

    prompt = """
Please analyze the attached PDF page image and identify all graphs present.

For each graph, I need a precise data extraction:

1. CHART TYPE IDENTIFICATION:
   - Identify the specific chart type (e.g., Line, Bar, Stacked Bar, Combo Chart with bars+lines, Bubble, Scatter, Pie)
   - For combo charts, explicitly identify which data series are bars and which are lines
   - Note any secondary Y-axis and which series use it

2. DATA SERIES EXTRACTION - WITH SPECIAL ATTENTION TO LINE GRAPHS:
   - For LINE GRAPHS:
     * Extract ALL significant points including peaks, valleys, and inflection points
     * Use fractional time points (e.g., 2017.25, 2017.5, 2017.75) when necessary to capture quarterly or monthly data
     * Capture at least 30-40 data points for complex lines with multiple rises and falls
     * Include ALL local maxima and minima, not just year-end values

   - For BAR GRAPHS:
     * Measure the exact height of each bar
     * For stacked bars: provide values for each segment separately

   - For COMBO CHARTS:
     * Extract line data with the same detail as pure line charts
     * Clearly distinguish which data belongs to bars vs. lines

3. AXIS INFORMATION:
   - X-axis: all labels and scale type (linear, logarithmic, categorical)
   - Y-axis: range, scale, units, and tick intervals
   - Include any secondary Y-axis details

4. LEGEND INTERPRETATION:
   - Map each visual element (bar/line/area) to its correct legend entry
   - Note color, pattern, or other distinguishing features
   - Maintain the exact naming from the legend

5. OUTPUT FORMAT:
   - For LINE GRAPHS with time series: use fractional time values in CSV (e.g., 2017, 2017.25, 2017.5, 2017.75)
   - For LINE GRAPHS with complex curves: provide enough points to accurately recreate the curve (30+ points)
   - Include the chart type, title, and axes information
   - Include a bounding box for the chart area as {left, top, width, height}

Be extremely precise with measurements. Extrapolate data points based on gridlines and axis scales. When in doubt about exact values, provide your best estimate based on visual proportions.

Return ONLY valid JSON in the following format with no other text before or after:
{
  "graphs": [
    {
      "graph_number": 1,
      "chart_type": "Line Chart",
      "data_representation": {"Line series": ["Current", "Avg", "+1SD", "+2SD", "-1SD", "-2SD"]},
      "csv_data": "Year,Current,Avg,+1SD,+2SD,-1SD,-2SD\\n2017,35.0,38.4,50.7,62.9,26.2,14.0\\n2017.25,36.2,38.4,50.7,62.9,26.2,14.0\\n2017.5,38.1,38.4,50.7,62.9,26.2,14.0\\n...",
      "axes_info": {
        "x_axis": {"label": "Year", "type": "continuous"},
        "y_axis": {"label": "P/E", "min": 0, "max": 80, "interval": 20}
      },
      "bounding_box": {"left": 100, "top": 150, "width": 400, "height": 300}
    }
  ]
}
"""

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-3-7-sonnet-20250219",
            max_tokens=4096,
            temperature=0.0,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": base64_encoded_image}}
                ]
            }]
        )
        raw_response = response.content[0].text
        debug_print("Raw response from Claude for", image_path, ":", raw_response)

        # Improved JSON extraction
        try:
            # First, try direct JSON parsing
            result_json = json.loads(raw_response)
        except json.JSONDecodeError:
            # If that fails, try to extract JSON content using regex
            import re
            json_pattern = r'({[\s\S]*})'
            json_match = re.search(json_pattern, raw_response)

            if json_match:
                json_str = json_match.group(1)
                try:
                    # Clean up potential trailing commas which are invalid in JSON
                    json_str = re.sub(r',\s*}', '}', json_str)
                    json_str = re.sub(r',\s*]', ']', json_str)
                    # Remove comments
                    json_str = re.sub(r'//.*', '', json_str)
                    result_json = json.loads(json_str)
                except json.JSONDecodeError as e:
                    print(f"Error parsing extracted JSON: {e}")
                    print(f"Attempted to parse: {json_str[:100]}...")
                    return []
            else:
                print("No JSON content found in Claude's response")
                return []

        graphs = result_json.get("graphs", [])

        # Additional processing for line charts to ensure detailed data points
        for graph in graphs:
            if "line" in graph.get("chart_type", "").lower():
                # Check if the CSV data has enough detail
                csv_data = graph.get("csv_data", "")
                rows = csv_data.count('\n') + 1

                # Log detailed info about line graph extraction
                debug_print(f"Line graph #{graph.get('graph_number', '?')} extracted with {rows} data points")

        # Filter out reference lines from each graph
        filtered_graphs = []
        for graph in graphs:
            filtered_graph = filter_out_reference_lines(graph)
            filtered_graphs.append(filtered_graph)

        debug_print("Detected graphs in", image_path, ":", filtered_graphs)
        return filtered_graphs

    except Exception as e:
        print("Error in detect_graphs_with_claude for", image_path, ":", e)
        return []


def extract_chart_from_pdf_precise(pdf_path, page_num, bbox, output_path, add_margin=5):
    """
    Extracts a chart region with precise adjustments for better accuracy.
    """
    import fitz  # PyMuPDF

    try:
        # Open the PDF document
        doc = fitz.open(pdf_path)

        # Convert from 1-based to 0-based page numbering
        page_idx = page_num - 1
        if page_idx < 0 or page_idx >= len(doc):
            debug_print(f"Page {page_num} out of range (document has {len(doc)} pages)")
            return None

        page = doc[page_idx]

        # Get PDF page dimensions
        page_width = page.rect.width
        page_height = page.rect.height
        debug_print(f"PDF page dimensions: {page_width} x {page_height}")

        # Get bbox coordinates
        left = bbox.get("left", 0)
        top = bbox.get("top", 0)
        width = bbox.get("width", 0)
        height = bbox.get("height", 0)

        debug_print(f"Original bbox: left={left}, top={top}, width={width}, height={height}")

        # CORRECTION FACTORS:
        # Based on your image, it appears the bounding boxes are far too large
        # Let's try applying a scaling factor to make them more precise

        # 1. Try a percentage of the original size (focused on center)
        scale_factor = 0.5  # Try extracting 50% of the original bbox size
        center_x = left + (width / 2)
        center_y = top + (height / 2)

        new_width = width * scale_factor
        new_height = height * scale_factor

        new_left = center_x - (new_width / 2)
        new_top = center_y - (new_height / 2)

        # Create rect with the adjusted coordinates
        rect = fitz.Rect(
            new_left - add_margin,
            new_top - add_margin,
            new_left + new_width + add_margin,
            new_top + new_height + add_margin
        )

        # Make sure the rect is within page bounds
        page_rect = page.rect
        rect = rect.intersect(page_rect)

        debug_print(f"Adjusted extraction rect: {rect}")

        # Extract the image with higher resolution
        pix = page.get_pixmap(clip=rect, matrix=fitz.Matrix(3, 3))  # 3x scaling for better resolution
        pix.save(output_path)

        debug_print(f"Extracted chart from PDF page {page_num} to {output_path}")

        return output_path

    except Exception as e:
        debug_print(f"Error extracting chart from PDF: {e}")
        return None
    finally:
        if 'doc' in locals():
            doc.close()

#############################################
# CSV-to-DataFrame Conversion
#############################################
def convert_csv_to_dataframe(csv_text):
    """
    Converts CSV text into a pandas DataFrame.
    """
    try:
        csv_text = csv_text.strip()
        return pd.read_csv(io.StringIO(csv_text))
    except Exception as e:
        print(f"Error converting CSV text to DataFrame: {e}")
        return None

#############################################
# Original Graph Cropping using pdfplumber
#############################################

def validate_crop(cropped_img, min_size=100):
    """
    Validates that a cropped image is likely to contain a valid chart.

    Args:
        cropped_img: PIL Image of the cropped chart
        min_size: Minimum width/height to be considered valid

    Returns:
        Boolean indicating if the crop is valid
    """
    if cropped_img is None:
        return False

    width, height = cropped_img.size
    if width < min_size or height < min_size:
        return False

    # You could add more validation here, like checking for
    # color variance to ensure it's not just a blank area

    return True


#############################################
# Excel Tab Creation (CSV Data & Original Graph)
#############################################
def create_excel_tab_with_data(dataframe, excel_writer, tab_name, chart_type, y_axis_info, secondary_y_axis, orig_img):
    """
    Creates a new sheet in the Excel file with the extracted DataFrame,
    plus the original cropped graph image on the right side.
    """
    try:
        dataframe.to_excel(excel_writer, sheet_name=tab_name, index=False, startrow=1)
        worksheet = excel_writer.sheets[tab_name]
        worksheet['A1'] = "Graph Analysis Results"
        worksheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)

        # Store the chart metadata in "Z1", "Z2", "Z3"
        worksheet["Z1"] = chart_type
        worksheet["Z2"] = y_axis_info
        worksheet["Z3"] = secondary_y_axis

        num_data_rows = len(dataframe) + 1
        worksheet["AA1"] = num_data_rows  # We'll use this to place the chart below the table

        gap_row = num_data_rows + 3
        worksheet["J" + str(gap_row - 1)] = "NOTE: This program is a work in progress; graph cropping may be imperfect."

        # Insert the original graph image
        if orig_img is not None:
            new_width = 600
            new_height = int(new_width * orig_img.height / orig_img.width)
            resized_image = orig_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            img_stream = io.BytesIO()
            resized_image.save(img_stream, format="PNG")
            img_stream.seek(0)
            worksheet.add_image(XLImage(img_stream), "J" + str(gap_row))
    except Exception as e:
        print(f"Error creating tab {tab_name} in Excel: {e}")

#############################################
# Chart Insertion Functions (Generated Charts)
#############################################
def map_chart_type(chart_type_val):
    """
    Maps a textual chart type to Excel's numeric chart constants via xlwings.
    Additional types can be added here as needed.
    """
    if not chart_type_val:
        return None, False

    ct = chart_type_val.lower()
    # If the user specifically says "bubble"
    if "bubble" in ct:
        return 15, False  # 15 = xlBubble
    # If the user says "scatter"
    if "scatter" in ct:
        return 74, False  # 74 = xlXYScatterLines
    # If the user says "and" or "combo" => combination bar+line
    if "and" in ct or "combo" in ct:
        return 51, True  # 51 = xlColumnClustered, True => combo
    # If the user says "stacked"
    if "stacked" in ct:
        # If you want vertical stacked columns, use 52 = xlColumnStacked
        return 52, False
    # If the user says "3d"
    if "3d" in ct:
        if "bar" in ct:
            return 15, False  # or 60 = xl3DBarClustered
        elif "pie" in ct:
            return -4160, False  # xl3DPie
        elif "line" in ct:
            return 65, False  # xl3DLine
        else:
            return None, False
    # Standard "pie"
    if "pie" in ct:
        return -4169, False  # xlPie
    # Standard "bar"
    if "bar" in ct:
        return 51, False  # xlColumnClustered
    # Standard "line"
    if "line" in ct:
        return 65, False  # xlLine
    # Otherwise, unknown
    return None, False

def nice_scale(data):
    """
    Sets a "nice" scale for the Y-axis based on data distribution.
    """
    import math
    import numpy as np
    data_array = np.array(data, dtype=float)
    raw_min = data_array.min()
    raw_max = data_array.max()
    data_range = raw_max - raw_min
    if data_range <= 0:
        return raw_min, raw_max, 1
    step = data_range / 5
    power = 10 ** math.floor(math.log10(step))
    step = math.ceil(step / power) * power
    new_min = math.floor(raw_min / step) * step
    new_max = math.ceil(raw_max / step) * step
    return new_min, new_max, step

def insert_nonpie_charts_xlwings(excel_output_path):
    """
    Inserts non-pie charts (including bubble, scatter, stacked bar, line, bar, combos) using xlwings.
    """
    try:
        app = xw.App(visible=False)
        wb = xw.Book(excel_output_path)

        for sheet in wb.sheets:
            chart_type_val = sheet.range("Z1").value
            # Skip pie charts – they are handled separately in insert_pie_charts_openpyxl.
            if chart_type_val and "pie" in chart_type_val.lower():
                continue

            data_range = sheet.range("A2").expand()
            if data_range.shape[0] < 2 or data_range.shape[1] < 2:
                continue

            nrows, ncols = data_range.shape
            y_axis_val = sheet.range("Z2").value
            secondary_y_axis_val = sheet.range("Z3").value
            table_end = int(sheet["AA1"].value)
            left = sheet.range("A" + str(table_end + 2)).left
            top = sheet.range("A" + str(table_end + 2)).top
            width = 350
            height = 210

            chart_obj = sheet.api.ChartObjects().Add(left, top, width, height)
            chart_com = chart_obj.Chart

            excel_chart_type, combo = map_chart_type(chart_type_val)
            # If chart type is unsupported, skip chart insertion.
            if excel_chart_type is None:
                debug_print(f"Unsupported chart type '{chart_type_val}' for sheet '{sheet.name}'. Skipping chart insertion.")
                continue

            # Special handling for line charts with fractional time points
            is_line_chart = "line" in chart_type_val.lower()
            has_fractional_time = False

            if is_line_chart:
                # Check if first column contains fractional time values
                time_col = sheet.range("A3").expand("down").value
                if isinstance(time_col, list):
                    for val in time_col:
                        if isinstance(val, (int, float)) and val % 1 != 0:
                            has_fractional_time = True
                            break
                elif isinstance(time_col, (int, float)) and time_col % 1 != 0:
                    has_fractional_time = True

            # For line charts with fractional time, use XY Scatter Lines instead of regular line chart
            if is_line_chart and has_fractional_time:
                excel_chart_type = 74  # xlXYScatterLines
                debug_print(f"Using XY Scatter for line chart with fractional time points on sheet '{sheet.name}'")

            # Handle special logic for bubble chart: we need X, Y, and bubble size.
            if "bubble" in chart_type_val.lower():
                # For bubble, let's assume first col = X, second col = Y, third col = size
                # If you have multiple bubble series, you can adapt this logic.
                if ncols < 3:
                    debug_print(f"Not enough columns for bubble chart on sheet '{sheet.name}'. Need at least 3.")
                    continue
                # Create a bubble chart
                chart_com.ChartType = excel_chart_type  # 15 = xlBubble
                # We treat each row (after the header) as one bubble data point.
                # In Excel's object model, each series can hold multiple bubbles, or we can treat each row as a separate series.
                # For simplicity, let's treat the entire table as one bubble series.
                # Set up the X, Y, and bubble size references
                x_vals = sheet.range("A3").expand("down").api
                y_vals = sheet.range("B3").expand("down").api
                size_vals = sheet.range("C3").expand("down").api
                series = chart_com.SeriesCollection().NewSeries()
                series.XValues = x_vals
                series.Values = y_vals
                series.BubbleSizes = size_vals
                # Optional: set a series name from the second row?
                series.Name = "Bubble Series"
                chart_com.HasTitle = True
                chart_com.ChartTitle.Text = "Recreated Bubble Chart"
                # Attempt axis labeling
                chart_com.Axes(1).HasTitle = True
                chart_com.Axes(1).AxisTitle.Text = "X Axis"
                chart_com.Axes(2).HasTitle = True
                if y_axis_val:
                    try:
                        if isinstance(y_axis_val, str):
                            parts = y_axis_val.split(",")
                            y_label = parts[0].strip() if parts else "Value"
                        elif isinstance(y_axis_val, dict):
                            y_label = y_axis_val.get("label", "Value")
                        else:
                            y_label = "Value"
                    except Exception as e:
                        debug_print(f"Error parsing Y-axis info on sheet {sheet.name}: {e}")
                        y_label = "Value"
                else:
                    y_label = "Value"
                chart_com.Axes(2).AxisTitle.Text = y_label
                continue  # Done handling bubble chart for this sheet

            # For other chart types, we do the normal logic:
            # If we have more rows than columns, we treat each row as a series.
            # If we have more columns than rows, we treat each column as a series.
            if (nrows - 1) < (ncols - 1):
                # Row-based data
                cat_range = sheet.range(data_range.address).offset(0, 0).resize(1, ncols)
                ser_range = sheet.range(data_range.address).offset(1, 0).resize(nrows - 1, ncols)
                chart_com.SetSourceData(Source=ser_range.api)
                chart_com.PlotBy = 1  # xlRows
                sc = chart_com.SeriesCollection()
                for i in range(1, sc.Count + 1):
                    sc.Item(i).XValues = cat_range.api
                    series_name = sheet.range("A2").offset(i, 0).value
                    sc.Item(i).Name = series_name
            else:
                # Column-based data
                cat_range = sheet.range(data_range.address).offset(1, 0).resize(nrows - 1, 1)
                ser_range = sheet.range(data_range.address).offset(1, 1).resize(nrows - 1, ncols - 1)
                chart_com.SetSourceData(Source=ser_range.api)
                chart_com.PlotBy = 2  # xlColumns
                sc = chart_com.SeriesCollection()
                for i in range(1, sc.Count + 1):
                    sc.Item(i).XValues = cat_range.api
                    # Try naming from the top row
                    series_name = sheet.range("B2").offset(0, i - 1).value
                    sc.Item(i).Name = series_name

            # Set the chart type
            chart_com.ChartType = excel_chart_type
            chart_com.HasTitle = True
            chart_com.ChartTitle.Text = "Recreated Graph"

            # If line or 3D line, disable 3D if needed
            if excel_chart_type == 65:  # xlLine
                try:
                    chart_com.ChartArea.Format.ThreeD.Visible = False
                    chart_com.PlotArea.Format.ThreeD.Visible = False
                    chart_com.Rotation = 0
                    chart_com.Elevation = 0
                except Exception as e:
                    debug_print("Error disabling 3D for line chart:", e)

            # For detailed line charts, disable markers if there are many points
            if is_line_chart and nrows > 15:
                sc = chart_com.SeriesCollection()
                for i in range(1, sc.Count + 1):
                    try:
                        # Make the lines smoother
                        sc.Item(i).MarkerStyle = -4142  # No marker
                        sc.Item(i).Smooth = True
                    except Exception as e:
                        debug_print(f"Error setting line style for series {i}: {e}")

            # Try labeling the X and Y axes if not pie or 3D pie
            if excel_chart_type not in [-4169, -4160]:
                # X axis label
                if (nrows - 1) < (ncols - 1):
                    x_title = sheet.range(data_range.address).offset(0, 0).value
                else:
                    x_title = sheet.range("A2").value
                chart_com.Axes(1).HasTitle = True
                chart_com.Axes(1).AxisTitle.Text = x_title if x_title else "X Axis"

                # Y axis label
                chart_com.Axes(2).HasTitle = True
                if y_axis_val:
                    try:
                        if isinstance(y_axis_val, str):
                            parts = y_axis_val.split(",")
                            y_label = parts[0].strip() if parts else "Value"
                        elif isinstance(y_axis_val, dict):
                            y_label = y_axis_val.get("label", "Value")
                        else:
                            y_label = "Value"
                    except Exception as e:
                        print(f"Error parsing Y-axis info on sheet {sheet.name}: {e}")
                        y_label = "Value"
                else:
                    y_label = "Value"
                chart_com.Axes(2).AxisTitle.Text = y_label

                # Attempt a "nice" scale
                data_values = ser_range.value
                numeric_values = []
                if isinstance(data_values, list):
                    for row in data_values:
                        if isinstance(row, (list, tuple)):
                            for cell in row:
                                try:
                                    numeric_values.append(float(cell))
                                except Exception:
                                    pass
                        else:
                            try:
                                numeric_values.append(float(row))
                            except Exception:
                                pass
                if numeric_values:
                    new_min, new_max, major_unit = nice_scale(numeric_values)
                    try:
                        chart_com.Axes(2).MinimumScale = float(new_min)
                        chart_com.Axes(2).MaximumScale = float(new_max)
                        chart_com.Axes(2).MajorUnit = float(major_unit)
                    except Exception as e:
                        print(f"Error setting Y-axis scaling on sheet {sheet.name}: {e}")

            # Secondary axis handling
            if str(secondary_y_axis_val).lower() == "true":
                try:
                    series_count = chart_com.SeriesCollection().Count
                    # Example: we place series 2..N on the secondary axis
                    for i in range(2, series_count + 1):
                        chart_com.SeriesCollection(i).AxisGroup = 2
                    debug_print("Set series 2 to", series_count, "to secondary axis.")
                except Exception as e:
                    debug_print("Error setting secondary axis on sheet", sheet.name, ":", e)

            # Combination (Bar and Line) => set last series to line
            if combo:
                try:
                    series_count = chart_com.SeriesCollection().Count
                    # The last series => line
                    chart_com.SeriesCollection(series_count).ChartType = 65  # xlLine
                    debug_print("Adjusted series", series_count, "to line chart for combo chart.")
                except Exception as e:
                    debug_print("Error adjusting combo chart series:", e)

        wb.save()
        wb.close()
        app.quit()
        del wb
        del app
        time.sleep(2)
        print(f"Non-pie charts inserted using xlwings in {excel_output_path}")

    except Exception as e:
        print(f"Error inserting non-pie charts with xlwings: {e}")

def insert_pie_charts_openpyxl(excel_output_path):
    """
    Inserts pie charts using openpyxl for sheets that have chart_type with "pie".
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.label import DataLabelList

        wb = load_workbook(excel_output_path)
        for sheet in wb.worksheets:
            chart_type_cell = sheet["Z1"].value
            if chart_type_cell and "pie" in chart_type_cell.lower():
                table_end_row = sheet["AA1"].value
                if not table_end_row or int(table_end_row) < 2:
                    continue
                last_row = int(table_end_row)
                data = Reference(sheet, min_col=2, min_row=3, max_row=last_row)
                cats = Reference(sheet, min_col=1, min_row=3, max_row=last_row)
                pie = PieChart()
                pie.title = "Recreated Graph"
                pie.add_data(data, titles_from_data=False)
                pie.set_categories(cats)
                dlabels = DataLabelList()
                dlabels.showPercent = False
                dlabels.showVal = True
                dlabels.showCatName = False
                dlabels.showSerName = False
                dlabels.number_format = "General"
                pie.dataLabels = dlabels
                chart_position = "A" + str(last_row + 2)
                sheet.add_chart(pie, chart_position)

        try:
            wb.save(excel_output_path)
        except PermissionError:
            print("PermissionError: Please close the file", excel_output_path, "and try again.")
        print(f"Pie charts inserted using openpyxl in {excel_output_path}")

    except Exception as e:
        print(f"Error inserting pie charts with openpyxl: {e}")

#############################################
# Process Integration
#############################################
def process_pdf_to_excel(pdf_path, excel_output_path, selected_pages=None, progress_callback=None):
    """
    Main function integrating Upstage and Claude:
    1. Reads API keys
    2. Calls Upstage to identify figures/charts in the PDF
    3. For each identified figure:
       a. Extract the figure directly from the PDF using PyMuPDF
       b. Call Claude with the image to get detailed data extraction
    4. Create Excel file with data, original image, and generated charts
    """
    try:
        # Read API keys
        anthropic_key = read_anthropic_key()
        upstage_key = read_upstage_api_key()

        if not anthropic_key:
            print("Could not authenticate with Claude API.")
            return

        if not upstage_key:
            print("Could not authenticate with Upstage API.")
            print("Falling back to direct PDF processing without Upstage...")
            return process_pdf_without_upstage(pdf_path, excel_output_path, anthropic_key)

        # Step 1: Call Upstage to identify figures in the PDF
        print("Calling Upstage Document Parser API...")
        if progress_callback:
            progress_callback("Identifying charts in PDF with Upstage...", 10)

        upstage_response = extract_figures_with_upstage(pdf_path, upstage_key)
        if not upstage_response:
            print("Upstage extraction failed. Falling back to direct PDF processing...")
            return process_pdf_without_upstage(pdf_path, excel_output_path, anthropic_key)

        # Process Upstage output
        figures = process_upstage_output(upstage_response, pdf_path)

        if not figures:
            print("No figures identified by Upstage. Falling back to direct PDF processing...")
            return process_pdf_without_upstage(pdf_path, excel_output_path, anthropic_key)

        print(f"Upstage identified {len(figures)} potential charts/figures")

        # Step 2: Process each figure individually
        if progress_callback:
            progress_callback("Analyzing charts with Claude AI...", 30)

        # Initialize Excel writer
        excel_writer = pd.ExcelWriter(excel_output_path, engine="openpyxl")
        sheet_added = False

        # Create output directories
        os.makedirs("extracted_charts", exist_ok=True)

        # Process each figure from Upstage
        for idx, fig in enumerate(figures):
            page_num = fig.get("page", 1)

            # Skip if page not in selected pages
            if selected_pages and page_num not in selected_pages:
                continue

            debug_print(f"Processing figure {idx+1}/{len(figures)} from page {page_num}")
            if progress_callback:
                progress_callback(f"Processing figure {idx+1} of {len(figures)}...",
                                  30 + (idx * 60 // len(figures)))

            # Get bounding box
            bbox = fig.get("bounding_box")
            if not bbox:
                debug_print(f"No bounding box for figure {idx+1}, skipping")
                continue

            # Extract the chart directly from PDF
            chart_path = os.path.join("extracted_charts", f"figure_{idx+1}.png")
            extracted_path = extract_chart_from_pdf(pdf_path, page_num, bbox, chart_path)

            if not extracted_path:
                debug_print(f"Failed to extract figure {idx+1} from PDF, skipping")
                continue

            # Load the extracted image for later use
            try:
                cropped_img = Image.open(extracted_path)
            except Exception as e:
                debug_print(f"Failed to load extracted image {extracted_path}: {e}")
                cropped_img = None

            # Process with Claude
            graphs = detect_graphs_with_claude(extracted_path, anthropic_key)

            if not graphs or len(graphs) == 0:
                debug_print(f"Claude found no graphs in figure {idx+1}, creating minimal tab")

                # Create a tab with minimal data but include the cropped image
                df = pd.DataFrame({"Note": ["No data could be extracted automatically"]})
                tab_name = f"Figure_{idx+1}"

                create_excel_tab_with_data(
                    df, excel_writer, tab_name, "Unknown", "", "False", cropped_img)
                sheet_added = True
            else:
                # Process each graph Claude found
                for g_idx, graph in enumerate(graphs):
                    csv_text = graph.get("csv_data")
                    if not csv_text:
                        debug_print(f"No CSV data for graph {g_idx+1} in figure {idx+1}")
                        continue

                    # Convert to DataFrame
                    df = convert_csv_to_dataframe(csv_text)
                    if df is None or df.empty:
                        debug_print(f"Empty DataFrame for graph {g_idx+1} in figure {idx+1}")
                        continue

                    # Get chart info
                    chart_type = graph.get("chart_type", "Unknown")

                    # Get Y-axis info
                    if "axes_info" in graph and "y_axis" in graph["axes_info"]:
                        y_axis = graph["axes_info"]["y_axis"]
                        if isinstance(y_axis, dict):
                            y_label = y_axis.get("label", "")
                            y_min = y_axis.get("min", "")
                            y_max = y_axis.get("max", "")
                            y_interval = y_axis.get("interval", "")
                            y_axis_info = f"{y_label},{y_min},{y_max},{y_interval}"
                        else:
                            y_axis_info = str(y_axis)
                    else:
                        y_axis_info = ""

                    # Check for secondary Y-axis
                    secondary_y_axis = "False"
                    if "axes_info" in graph and "secondary_y_axis" in graph["axes_info"]:
                        secondary_y_axis = "True"
                    elif "axes_info" in graph and "y_axis_secondary" in graph["axes_info"]:
                        secondary_y_axis = "True"

                    # Generate tab name
                    tab_name = f"Figure_{idx+1}"
                    if g_idx > 0:  # If Claude found multiple graphs in one figure
                        tab_name += f"_{g_idx+1}"

                    # Create Excel tab
                    create_excel_tab_with_data(
                        df, excel_writer, tab_name, chart_type,
                        y_axis_info, secondary_y_axis, cropped_img)
                    sheet_added = True

        # Finalize the Excel file
        if sheet_added:
            excel_writer.close()
            print(f"Excel file created: {excel_output_path}")

            if progress_callback:
                progress_callback("Generating charts in Excel...", 90)

            # Clean up any temporary Excel files
            remove_temp_excel_files(excel_output_path)
            wait_for_file_lock_release(excel_output_path)

            # Add charts to the Excel file
            insert_nonpie_charts_xlwings(excel_output_path)
            time.sleep(2)
            wait_for_file_lock_release(excel_output_path)
            insert_pie_charts_openpyxl(excel_output_path)

            if progress_callback:
                progress_callback("Processing complete!", 100)
        else:
            print("No valid data extracted. Skipping Excel file creation.")

    except Exception as e:
        print(f"Error processing PDF {pdf_path}: {e}")
        import traceback
        traceback.print_exc()

def create_pdf_extraction_debug(pdf_path, figures, output_dir="debug_extractions"):
    """
    Creates debug visualizations of chart extraction regions.

    Args:
        pdf_path: Path to the PDF file
        figures: List of figure dictionaries from Upstage
        output_dir: Directory to save debug images
    """
    try:
        import fitz
        import cv2
        import numpy as np

        os.makedirs(output_dir, exist_ok=True)
        doc = fitz.open(pdf_path)

        for idx, fig in enumerate(figures):
            page_num = fig.get("page", 1)
            page_idx = page_num - 1

            if page_idx < 0 or page_idx >= len(doc):
                continue

            page = doc[page_idx]
            bbox = fig.get("bounding_box")

            if not bbox:
                continue

            # Render the whole page
            pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            img_data = pix.samples
            img = np.frombuffer(img_data, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)

            if pix.n == 4:  # RGBA
                img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)

            # Draw a rectangle around the figure
            left = bbox.get("left", 0) * 1.5  # Scale to match the matrix we used
            top = bbox.get("top", 0) * 1.5
            width = bbox.get("width", 0) * 1.5
            height = bbox.get("height", 0) * 1.5

            cv2.rectangle(img,
                          (int(left), int(top)),
                          (int(left + width), int(top + height)),
                          (0, 255, 0), 3)

            # Add label
            cv2.putText(img, f"Figure {idx+1}", (int(left), int(top - 10)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

            # Save the debug image
            debug_path = os.path.join(output_dir, f"page_{page_num}_figure_{idx+1}.png")
            cv2.imwrite(debug_path, img)

            # Also extract just the figure region
            rect = fitz.Rect(
                bbox.get("left", 0),
                bbox.get("top", 0),
                bbox.get("left", 0) + bbox.get("width", 0),
                bbox.get("top", 0) + bbox.get("height", 0)
            )

            crop_pix = page.get_pixmap(clip=rect, matrix=fitz.Matrix(2, 2))
            crop_path = os.path.join(output_dir, f"page_{page_num}_figure_{idx+1}_crop.png")
            crop_pix.save(crop_path)

        doc.close()
        print(f"Debug extractions saved to {output_dir}")

    except Exception as e:
        debug_print(f"Error creating extraction debug: {e}")

def get_accurate_pdf_dimensions(pdf_path):
    """Get actual PDF dimensions using pdfplumber."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            first_page = pdf.pages[0]
            # Get the media box (page size)
            if hasattr(first_page, 'mediabox'):
                pdf_width = first_page.mediabox[2] - first_page.mediabox[0]
                pdf_height = first_page.mediabox[3] - first_page.mediabox[1]
            else:
                # Fallback
                pdf_width = first_page.width
                pdf_height = first_page.height

            # Check orientation
            is_landscape = pdf_width > pdf_height

            return pdf_width, pdf_height, is_landscape
    except Exception as e:
        debug_print(f"Error getting PDF dimensions: {e}")
        # Default to standard letter
        return 612, 792, False




def process_pdf_without_upstage(pdf_path, excel_output_path, anthropic_key):
    """
    Fallback function that processes PDF without Upstage API:
    1. Extracts each page from the PDF using PyMuPDF
    2. Sends each page directly to Claude for analysis
    """
    print("Processing PDF without Upstage...")
    try:
        import fitz  # PyMuPDF

        # Create output directory
        os.makedirs("extracted_pages", exist_ok=True)

        # Extract all pages from the PDF
        doc = fitz.open(pdf_path)
        page_images = []

        for page_idx in range(len(doc)):
            page = doc[page_idx]
            page_num = page_idx + 1  # 1-based page numbering

            # Render the page to an image
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x scaling for better quality
            img_path = os.path.join("extracted_pages", f"page_{page_num}.png")
            pix.save(img_path)
            page_images.append(img_path)
            debug_print(f"Extracted page {page_num} saved to {img_path}")

        doc.close()

        all_graphs = []

        for page_image in page_images:
            print(f"Detecting graphs in {page_image}...")
            graphs = detect_graphs_with_claude(page_image, anthropic_key)
            if graphs:
                for graph in graphs:
                    all_graphs.append((graph, page_image))
            else:
                debug_print(f"No graphs detected in {page_image}.")

        # Create Excel with openpyxl engine
        excel_writer = pd.ExcelWriter(excel_output_path, engine="openpyxl")
        sheet_added = False

        # Process graphs as normal
        for graph_data, page_image in all_graphs:
            csv_text = graph_data.get("csv_data")
            chart_type = graph_data.get("chart_type")

            # Handle different structures for axes info
            if "axes_info" in graph_data and "y_axis" in graph_data["axes_info"]:
                y_axis = graph_data["axes_info"]["y_axis"]
                if isinstance(y_axis, dict):
                    y_label = y_axis.get("label", "")
                    y_min = y_axis.get("min", "")
                    y_max = y_axis.get("max", "")
                    y_interval = y_axis.get("interval", "")
                    y_axis_info = f"{y_label},{y_min},{y_max},{y_interval}"
                else:
                    y_axis_info = str(y_axis)
            else:
                y_axis_info = graph_data.get("y_axis_info", "")

            # Check for secondary Y-axis
            secondary_y_axis = graph_data.get("secondary_y_axis", "False")

            if not csv_text:
                print(f"No CSV data for graph {graph_data.get('graph_number', '?')}.")
                continue

            # Convert CSV to DataFrame
            df = convert_csv_to_dataframe(csv_text)
            if df is None or df.empty:
                continue

            # Create tab name from page and graph number
            page_num = os.path.basename(page_image).split('_')[1].split('.')[0]
            graph_number = graph_data.get("graph_number", 1)
            tab_name = f"Page{page_num}_Graph{graph_number}"

            # Load the original page image for cropping
            try:
                page_img = Image.open(page_image)
                bbox = graph_data.get("bounding_box")

                if bbox:
                    left = bbox.get("left", 0)
                    top = bbox.get("top", 0)
                    width = bbox.get("width", 0)
                    height = bbox.get("height", 0)

                    # Crop the image
                    try:
                        orig_img = page_img.crop((left, top, left + width, top + height))
                    except Exception as e:
                        debug_print(f"Error cropping image: {e}")
                        orig_img = page_img  # Use full page as fallback
                else:
                    orig_img = page_img  # Use full page as fallback
            except Exception as e:
                debug_print(f"Error loading page image: {e}")
                orig_img = None

            # Create Excel tab
            create_excel_tab_with_data(
                df, excel_writer, tab_name, chart_type,
                y_axis_info, secondary_y_axis, orig_img
            )
            sheet_added = True

        if sheet_added:
            excel_writer.close()
            print(f"Excel file created: {excel_output_path}")

            # Clean up any temporary Excel files
            remove_temp_excel_files(excel_output_path)
            wait_for_file_lock_release(excel_output_path)

            # Add charts to the Excel file
            insert_nonpie_charts_xlwings(excel_output_path)
            time.sleep(2)
            wait_for_file_lock_release(excel_output_path)
            insert_pie_charts_openpyxl(excel_output_path)
        else:
            print("No valid data extracted. Skipping Excel file creation.")

    except Exception as e:
        print(f"Error in fallback processing: {e}")
        import traceback
        traceback.print_exc()

# Example usage
if __name__ == "__main__":
    pdf_file_path = r"C:\Users\Mbomm\IdeaProjects\PDF Graph Scanner\input_pdfs\Sample3.pdf"
    excel_file_path = "output_upstage.xlsx"

    # Simple function to display progress (replace with GUI progress bar in real app)
    def print_progress(message, percent):
        print(f"{message} - {percent}% complete")

    # Process with Upstage + Claude
    process_pdf_to_excel(pdf_file_path, excel_file_path, progress_callback=print_progress)